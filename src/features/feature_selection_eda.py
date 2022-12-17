import os
import re
import sys
import random
import time
import warnings
import lightgbm
import catboost
import openpyxl
import numpy as np
import pandas as pd
import seaborn as sns
import missingno as mn
from sklearn import svm
import matplotlib as mpl
from copy import deepcopy
from sklearn import tree
from sklearn import ensemble
from datetime import datetime
from collections import Counter
from scipy.fft import fft, ifft
import matplotlib.pyplot as plt
from sklearn import kernel_ridge
from scipy.stats import pearsonr
from sklearn import decomposition
from sklearn import linear_model
from sklearn import model_selection
from openpyxl import load_workbook
from dtaidistance import dtw, dtw_ndim
from sklearn.utils import class_weight
from sklearn.preprocessing import scale
from kneed import DataGenerator, KneeLocator
from sklearn.preprocessing import Normalizer
from sklearn.preprocessing import minmax_scale
from statsmodels.tsa.stattools import adfuller
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import RepeatedKFold
from IPython.display import display, Markdown, Image
from sklearn.model_selection import train_test_split
from sklearn.pipeline import make_pipeline, Pipeline
from sklearn.feature_selection import VarianceThreshold
from sklearn.metrics import mean_squared_error, mean_absolute_error
from fitter import Fitter, get_common_distributions, get_distributions
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.model_selection import KFold, cross_val_score, train_test_split
from sklearn.preprocessing import MinMaxScaler, StandardScaler, RobustScaler
from sklearn.metrics import classification_report, confusion_matrix, roc_curve
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error

CURRENT_FILE_PATH = os.path.dirname(os.path.abspath(__file__))
CURRENT_FILE_PATH_PARENT = os.path.dirname(CURRENT_FILE_PATH)
PRPOJECT_PATH = os.path.dirname(CURRENT_FILE_PATH_PARENT)
FEATURE_FILE = CURRENT_FILE_PATH + "/FeatureCSV.csv"

sys.path.append(CURRENT_FILE_PATH_PARENT)
from data.data_utility import DataUtility

# Feature Selection
class SelectingEDAFeatures(DataUtility):
    def __init__(self,
        parent_data_folder=None,
        raw_data_folder=None,
        zipped_folder=None,
        columns=None,
        flags="FD001",
        per_missing=0.50,
        dominant_threshold= 0.95,
        dispersion_threshold=(1-1.0e-9, 1+1.0e-9),
        heat_map_threshold = 0.9,
        max_vif=9,
        excel_file=None,
        feature_file = None,
        re_create_excel=True):
        """
        This module handles EDA and initial feature selection at the same time.
        EDA results are stored in an excel file

        Args:
            parent_data_folder (_type_, optional): _description_. Defaults to None.
            raw_data_folder (_type_, optional): _description_. Defaults to None.
            zipped_folder (_type_, optional): _description_. Defaults to None.
            columns (_type_, optional): _description_. Defaults to None.
            flags (str, optional): _description_. Defaults to "FD001".
            per_missing (float, optional): _description_. Defaults to 0.50.
            dominant_threshold (float, optional): _description_. Defaults to 0.95.
            dispersion_threshold (tuple, optional): _description_. Defaults to (1-1.0e-9, 1+1.0e-9).
            heat_map_threshold (float, optional): _description_. Defaults to 0.9.
            max_vif (int, optional): _description_. Defaults to 9.
            excel_file (_type_, optional): _description_. Defaults to None.
            re_create_excel (bool, optional): _description_. Defaults to True.
        """
        super().__init__(parent_data_folder=parent_data_folder,
                raw_data_folder=raw_data_folder,
                zipped_folder=zipped_folder,
                columns=columns)
        self.per_missing = per_missing
        self.dominant_threshold = dominant_threshold
        self.dispersion_threshold = dispersion_threshold
        self.all_feature_colums = self.train_columns
        self.heat_map_threshold = heat_map_threshold
        self.max_vif = max_vif
        self.DF_TRAIN, _, _= self.load_data_by_flags(flags=flags)

        self.report_folder = PRPOJECT_PATH
        if not excel_file:
            excel_file = self.report_folder + f"/reports/EDA_Reports{datetime.strftime(datetime.now(), '%Y%m%d%H')}.xlsx"
        self.excel_file = excel_file
        self.tmp_png = self.report_folder + f"/reports/tmp_{datetime.strftime(datetime.now(), '%Y%m%d%H')}.png"

        self.prepare_excel_report(re_create_excel=re_create_excel)

    # Excel
    def prepare_excel_report(self, re_create_excel=False):
        """
        Prepare Excel File for eda retports
        Args:
            re_create_excel (bool, optional): _description_. Defaults to False.
        """
        if re_create_excel or (not os.path.exists(self.excel_file)):
            if  os.path.exists(self.excel_file):
                os.remove(self.excel_file)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes"
            ws["A1"] = "About EDA"

            ws["A3"] = "File Locations:"
            n_row, n_col = 3, 1
            for k,v in self.file_folders.items():
                cell = ws.cell(n_row+1, n_col)
                cell.value = k
                cell = ws.cell(n_row+1, n_col+1)
                cell.value = v
                n_row += 1
            wb.save(self.excel_file)
            wb.close()

    # Excel
    def excel_add_data(self, data, sn, start_col, start_row, title):
        """
        Add data frame to excel sheet

        Args:
            data (_type_): _description_
            sn (_type_): _description_
            start_col (_type_): _description_
            start_row (_type_): _description_
            title (_type_): _description_
        """
        wb = load_workbook(self.excel_file)
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            ws.title=sn
        else:
            ws = wb[sn]
        ws.cell(column=start_col,row=start_row,value=title)

        cols = list(data)
        start_row += 1

        for i, col in enumerate(cols):
            ws.cell(column=start_col+i, row=start_row, value=col)

        start_row += 1
        vals = data.values
        nrow, ncol = vals.shape

        for r in range(nrow):
            for c in range(ncol):
                ws.cell(column=start_col+c, row=start_row+r, value=vals[r,c])
        wb.save(self.excel_file)
        wb.close()

    # Excel
    def excel_chart(self, sn, title, start_row=2):
        """
        Load chart to Excel

        Args:
            sn (_type_): _description_
            title (_type_): _description_
            start_row (int, optional): _description_. Defaults to 2.
        """
        wb = load_workbook(self.excel_file)
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            ws.title=sn
        else:
            ws = wb[sn]
        ws[f"A{start_row}"] = title

        img = openpyxl.drawing.image.Image(self.tmp_png)
        img.anchor = f"A{start_row+2}"

        ws.add_image(img)

        wb.save(self.excel_file)
        wb.close()

    # EDA
    def eda_describe(self):
        """
        This method output training data information to excel
        """
        tmp = self.DF_TRAIN.copy()
        agg = tmp.describe()

        sn = "Describe"
        start_row, start_col = 2, 1
        self.excel_add_data(agg, sn, start_col, start_row, "Describe Training Data")

        #few rows
        start_row += 20
        self.excel_add_data(tmp.sample(20), sn, start_col, start_row, "Sample Lines")

    # EDA
    def eda_plot_life(self, show_png = False):
        """
        plot engine life

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        plt.ioff()
        tmp = self.DF_TRAIN.copy()
        plt.style.use('seaborn')
        plt.figure(figsize=(15,20))

        ax=tmp.groupby('id')['cycle'].max().sort_values(ascending = True).plot(kind='barh',width=0.8, stacked=True,align='center',rot=0) #edge rot: rot=90
        plt.title('Engines Life Time',fontweight='bold',size=15)
        plt.xlabel('Cycles',fontweight='bold',size=12)
        plt.xticks(size=12)
        plt.ylabel('Engine ID',fontweight='bold',size=12)
        plt.yticks(size=12)

        for container in ax.containers:
            ax.bar_label(container, fontsize=12, color="blue", fmt='%g')

        plt.grid(False)
        plt.tight_layout()
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="Life", title="Engine Life Cycle")

        if show_png:
            plt.show()

        # distribution
        fig, (ax1, ax2)  = plt.subplots(1,2, sharex=True, figsize=(12,5))
        agg_df = tmp.groupby('id')['cycle'].max()
        agg_df.plot.hist(bins = 20, ax = ax1)
        agg_df = tmp.groupby('id')['cycle'].max()
        agg_df.plot.kde(ax = ax2)
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="Life", title="Engine Life Cycle Distribution", start_row=106)
        if show_png:
            plt.show()

    # EDA
    def eda_plot_missing(self, show_png=False):
        """
        Plot missing column information

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        plt.ioff()
        tmp = self.DF_TRAIN.copy()
        agg = tmp.isnull().sum().reset_index()
        agg.columns = ["Columns", "Missing#"]

        sn = "Missing"
        self.excel_add_data(agg, sn, 20, 2, "Missing Value Count by Column")

        plt.figure(figsize=(10, 8))
        mn.matrix(tmp, labels=True, sort="descending", figsize=(12,8))
        plt.savefig(self.tmp_png)
        self.excel_chart(sn=sn, title="Missing Information")

    # EDA
    def eda_hist_all_columns(self, show_png=False):
        """
        Histogram for all columns

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        plt.ioff()
        tmp = self.DF_TRAIN.copy()
        tmp.hist(bins=25, figsize=(15, 25), layout=(-1, 5), edgecolor="black")
        plt.tight_layout()
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="Hist", title="Histogram for all Columns", start_row=1)
        if show_png:
            plt.show()

    # EDA
    def eda_most_frequent(self, show_png=False):
        """
        Check how values count looks like for training data

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        plt.ioff()
        tmp = self.DF_TRAIN.copy()
        most_frequent_entry = tmp.mode()
        df_freq = tmp.eq(most_frequent_entry.values[0], axis=1)
        df_freq = df_freq.mean().sort_values(ascending=False)
        df_freq.plot.bar(figsize=(15, 4))
        plt.tight_layout()
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="Frequency", title="Most Frequent Value(Mode) Percentage", start_row=1)

        if show_png:
            plt.show()

    #EDA
    def eda_pair_plot(self, show_png=False, n_sensors=5, except_colmns=None):
        """
        Plot feature pairs. randomly take 5

        Args:
            show_png (bool, optional): _description_. Defaults to False.
            n_sensors (int, optional): _description_. Defaults to 5.
            except_colmns (_type_, optional): _description_. Defaults to None.
        """

        if not except_colmns:
            except_colmns = ["sensor1", "sensor5", "sensor10", "sensor16", "sensor18", "sensor19"]

        plt.ioff()
        id_engine = random.randint(1, 100)
        tmp = self.DF_TRAIN.copy()
        tmp = tmp[tmp["id"]==id].copy()

        sensors =[c for c in list(tmp) if ((re.search("sen", c)) and (c not in (except_colmns)))]

        sensors = random.sample(sensors, n_sensors)
        sensors.append("id")
        sensors.append("cycle")

        sns.pairplot(tmp[sensors], diag_kind = 'kde',   plot_kws={'alpha':0.5, 'edgecolor': 'r'})

        plt.savefig(self.tmp_png)
        self.excel_chart(sn="FeaturePair", title="Pair Plot for features", start_row=1)
        if show_png:
            plt.show()

    # EDA
    def eda_peer_grid_corner(self, show_png=False, n_sensors=5, except_colmns=None):
        """
        Peer plot. company to engines

        Args:
            show_png (bool, optional): _description_. Defaults to False.
            n_sensors (int, optional): _description_. Defaults to 5.
            except_colmns (_type_, optional): _description_. Defaults to None.
        """
        if not except_colmns:
            except_colmns = ["sensor1", "sensor5", "sensor10", "sensor16", "sensor18", "sensor19"]

        plt.ioff()
        id_engine1 = random.randint(1, 100)
        id_engine2 = random.randint(1, 100)

        tmp = self.DF_TRAIN.copy()

        ids = [id_engine1, id_engine2]

        tmp = tmp[tmp["id"].isin(ids)].copy()

        sensors =[c for c in list(tmp) if ((re.search("sen", c)) and (c not in (except_colmns)))]

        sensors = random.sample(sensors, n_sensors)
        sensors.append("id")
        sensors.append("cycle")
        sensors.append("rul")


        g = sns.PairGrid(tmp, vars = sensors,diag_sharey = False, corner = True, hue = 'id')
        g.map_lower(plt.scatter, alpha = 0.6)
        g.map_diag(plt.hist, alpha = 0.7)
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="PeerCompare", title="Compare two random engines", start_row=1)
        if show_png:
            plt.show()

    # EDA
    def eda_heatmap(self,  show_png=False):
        """
        EDA heatmap. pearson.

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        plt.ioff()
        tmp = self.DF_TRAIN.copy()
        df_corr = tmp.corr(method="pearson")
        labels = np.where(np.abs(df_corr)>0.75, "S",
                        np.where(np.abs(df_corr)>0.5, "M",
                                np.where(np.abs(df_corr)>0.25, "W", "")))

        plt.figure(figsize=(15, 15))
        sns.heatmap(df_corr, mask=np.eye(len(df_corr)), square=True,
                    center=0, annot=labels, fmt='', linewidths=.5,
                    cmap="vlag", cbar_kws={"shrink": 0.8});
        plt.savefig(self.tmp_png)
        self.excel_chart(sn="Heatmap", title="Heatmap for features", start_row=1)
        if show_png:
            plt.show()

    # EDA
    def eda(self,show_png=False):
        """
        EDA summary

        Args:
            show_png (bool, optional): _description_. Defaults to False.
        """
        self.eda_describe()
        self.eda_plot_life(show_png=show_png)
        self.eda_plot_missing(show_png=show_png)
        self.eda_hist_all_columns(show_png=show_png)
        self.eda_most_frequent(show_png=show_png)
        self.eda_pair_plot(show_png=show_png)
        self.eda_peer_grid_corner(show_png=show_png)
        self.eda_heatmap(show_png=show_png)

    def indentify_missing(self):
        """
            documentation:
            identify missing
        """
        tmp = self.DF_TRAIN.copy()
        # to test:
        #masks = np.random.choice([True, False], size=tmp.shape, p = [p, 1-p])
        #tmp = tmp.mask(masks)
        return np.array(list(tmp))[tmp.isnull().sum(axis=0)/tmp.shape[0] > self.per_missing]

    def dominant(self, x, per):
        """
            documentation:
            helper fun
        """
        return Counter(x).most_common(1)[0][1]/len(x) > per

    def identify_constants(self):
        """
            documentation:
            identify constants
        """
        tmp = self.DF_TRAIN.copy()
        tmp = tmp.select_dtypes(include = "number")
        tmp.drop(columns = ["id", "cycle", "rul"], inplace = True)

        feats =  list(tmp)

        # compare it in two way: threshold = 0
        vt = VarianceThreshold(threshold=0)
        _ = vt.fit_transform(tmp)
        kept_columns = vt.get_feature_names_out()

        consts = [c for c in feats if c not in kept_columns]

        # dominant values doesn't account for per%
        quasi_columns = np.array(list(tmp))[tmp.apply(lambda x: self.dominant(x, self.dominant_threshold))]

        return list(set(consts + list(quasi_columns)))

    def dispersion(self,data):
        """
            documentation:
            dispersions.
            return flat cols and dispersion values
        """
        arith_mean = np.mean(data+1, axis =0 )
        #geo_mean = np.power(np.prod(data, axis =0 ),1/data.shape[0])
        geo_mean = sum(np.log1p(data))/len(data)
        geo_mean = np.exp(geo_mean)
        return arith_mean/geo_mean

    def identify_flat_dispersion(self):
        tmp = self.DF_TRAIN.copy()
        tmp = tmp.select_dtypes(include = "number")
        tmp.drop(columns = ["id", "cycle", "rul"], inplace = True)

        dispersions = tmp.apply(lambda x: self.dispersion(x), axis=0)
        criteria = np.where(np.logical_and(dispersions>self.dispersion_threshold[0], \
                                           dispersions<self.dispersion_threshold[1]))
        low_dispersion = np.array(list(tmp))[criteria]
        return low_dispersion, dispersions
        #dispersions = np.array(list(tmp))[tmp.apply(lambda x: dispersion(x), axis=0)>threshold]
        #ROUNDONE_DISPERSION, _  = filter_by_dispersion(threshold=thresholds)

    def plot_by_id_corner(self, ids = [12,76], n_cols =5, pre_filtered_cols = None):
        """
            documentation:
            plotting corr
        """
        if not pre_filtered_cols:
            pre_filtered_cols = self.cols_for_plotting()

        tmp = self.DF_TRAIN.copy()
        tmp = tmp[tmp["id"].isin(ids)]

        max_rul = tmp["rul"].max()
        for id_engine in ids:
            if max_rul > tmp[tmp["id"]==id_engine]["rul"].max():
                max_rul = tmp[tmp["id"]==id_engine]["rul"].max()
        tmp = tmp[tmp["rul"]<=max_rul]

        cols = random.sample(pre_filtered_cols, n_cols)

        cols.append("rul")

        # seaborn matplotlib
        g = sns.PairGrid(tmp, vars = cols,diag_sharey = False, corner = True, hue = 'id' )
        g.map_lower(plt.scatter, alpha=0.6)
        #g.map_diag(plt.hist, alpha=0.7 )
        g.map_diag(sns.kdeplot)

        plt.show()
        #plot_by_id_corner()

    def plot_by_id_pair(self, ids = [2,76], n_cols =5, pre_filtered_cols = None):
        """
            documentation:
            plotting corr
        """
        sns.set_style("whitegrid")

        if not pre_filtered_cols:
            pre_filtered_cols = self.cols_for_plotting()

        tmp = self.DF_TRAIN.copy()
        tmp = tmp[tmp["id"].isin(ids)]

        max_rul = tmp["rul"].max()
        for id_engine in ids:
            if max_rul > tmp[tmp["id"]==id_engine]["rul"].max():
                max_rul = tmp[tmp["id"]==id_engine]["rul"].max()
        tmp = tmp[tmp["rul"]<=max_rul]

        cols = random.sample(pre_filtered_cols, n_cols)
        cols.append("rul")

        sns.pairplot(tmp, vars=cols, diag_kind = 'kde', hue="id")

        plt.show()
        #plot_by_id_pair()

    def plot_by_id_strip(self, ruls=None, n_feat=5, pre_filtered_cols=None):
        """
            documentation:
            plotting corr
        """
        sns.set_style("whitegrid")

        if not pre_filtered_cols:
            pre_filtered_cols = self.cols_for_plotting()

        tmp = self.DF_TRAIN.copy()

        if not ruls:
            ruls = [i for i in range(1, tmp["rul"].max(), 50)]

        tmp = tmp[tmp["rul"].isin(ruls)]

        cols = random.sample(pre_filtered_cols, n_feat)

        sns.stripplot(data=tmp, y="sensor21", x="rul",  dodge=True, jitter=False,  palette="deep")

        plt.show()
        return tmp

    def cols_for_plotting(self):
        """
            doc:
            helper function
        """
        constant_cols = self.identify_constants()
        return [c for c in self.all_feature_colums if c not in constant_cols]

    def heatmap_plot(self, show_fig=True, pre_filtered_cols=None ):
        """
            doc:
            return high corr cols
        """
        if not pre_filtered_cols:
            pre_filtered_cols = self.cols_for_plotting()

        tmp = self.DF_TRAIN.copy()
        tmp_corr = tmp[pre_filtered_cols].corr(method="pearson")

        if show_fig:
            plt.figure(figsize=(9, 9))

            ax = sns.heatmap(tmp_corr,  square=True, annot=True,annot_kws={"size": 8},
                        center=0, fmt=".2f",   linewidths=.5,  #fmt=".2g",
                        cmap="vlag", cbar_kws={"shrink": 0.8});

            ax.xaxis.tick_top()
            #sns.set(font_scale=1)
            plt.xticks(rotation = 90)
            plt.show()

        colss = list(tmp_corr)
        ln = len(tmp_corr)
        allvals = tmp_corr.values
        pairs = []
        for i in range(ln):
            for j in range(i+1, ln):
                if abs(allvals[i, j])>self.heat_map_threshold:
                    pairs.append((colss[i], colss[j], allvals[i, j]))
        return pairs

    def filter_by_val(self, feats):   # 9.0 is conventional chosen
        """
            doc:
            return high corr cols
        """
        ss = StandardScaler() # Normalizer, RobustScaler, minmax_scale
        tmp = self.DF_TRAIN.copy()
        tmp = tmp[feats]

        ss_tmp = ss.fit_transform(tmp)

        ss_tmp = pd.DataFrame(ss_tmp, columns=list(tmp))
        vif_data = pd.DataFrame()
        vif_data["feature"] = ss_tmp.columns

        vif_data["VIF"] = [variance_inflation_factor(ss_tmp.values, i) for i in range(len(ss_tmp.columns))]

        return list(vif_data["feature"]), vif_data

    def identify_high_vif_columns(self, pre_filtered_cols=None):
        """
            doc:
            return high corr cols from VIF
            return
        """
        if not pre_filtered_cols:
            pre_filtered_cols = self.cols_for_plotting()

        tmp = self.DF_TRAIN.copy()
        val = np.inf
        rounds = 0
        max_rounds = tmp.shape[1]

        feats = deepcopy(pre_filtered_cols)
        high_vif_col = []

        while val > self.max_vif:
            rounds += 1
            feats, df_vals = self.filter_by_val(feats=feats)

            max_row = df_vals.loc[df_vals["VIF"].idxmax()]

            feat, val = max_row[0], max_row[1]

            if val < self.max_vif:
                return list(feats), high_vif_col
            else:
                feats.remove(feat)
                high_vif_col.append(feat)
            if rounds>max_rounds:
                break
        return feats, high_vif_col

    def eda_features(self):
        """
            doc:
            output: conduct exploration and summarize
        """
        missing = self.indentify_missing()
        const = self.identify_constants()
        flats, _ = self.identify_flat_dispersion()
        _, vif = self.identify_high_vif_columns()

        to_remove = list(missing) + list(const) + list(flats) + list(vif)

        df1 = pd.DataFrame({"Feature": self.all_feature_colums})
        df2 =  pd.DataFrame({"Remove_Feature": to_remove})
        df1 = df1.merge(df2, left_on="Feature", right_on="Remove_Feature", how = "outer" )
        df1["Remove_Feature"] = df1["Remove_Feature"].fillna("(keep)")
        self.excel_add_data(df1, "Features", start_col =1, start_row=2, title="Feature Selection After EDA")

        df_ = pd.DataFrame({"Feature": [c for c in self.all_feature_colums if c not in to_remove]})
        df_.to_csv(FEATURE_FILE, index=False)

if __name__ =="__main__":
    fs = SelectingEDAFeatures()
    fs.eda_features()
    #fs.eda()
    #fs.prepare_excel_report()
    #print("done")