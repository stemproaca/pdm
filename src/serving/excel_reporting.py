
import os
import re
import sys
import random
import openpyxl
import numpy as np
import pandas as pd
import seaborn as sns
import missingno as mn
from sklearn import svm
import matplotlib as mpl
from copy import deepcopy
from sklearn import tree
from sklearn import ensemble
from datetime import datetime
from collections import Counter
from scipy.fft import fft, ifft
import matplotlib.pyplot as plt
from sklearn import kernel_ridge
from scipy.stats import pearsonr
from sklearn import decomposition
from sklearn import linear_model
from sklearn import model_selection
from openpyxl import load_workbook

CURRENT_FILE_PATH = os.path.dirname(os.path.abspath(__file__))
CURRENT_FILE_PATH_PARENT = os.path.dirname(CURRENT_FILE_PATH)
PRPOJECT_PATH = os.path.dirname(CURRENT_FILE_PATH_PARENT)
sys.path.append(CURRENT_FILE_PATH_PARENT)

from preparing.data_utility import DataUtility

# Feature Research. MRO
class ExcelUtility(DataUtility):
    def __init__(self,
        parent_data_folder=None,
        raw_data_folder=None,
        zipped_folder=None,
        excel_file_prefix = None,
        tmp_png_name_only = None,
        excel_file=None,
        re_create_excel=True):
        """
        This module handles Excel Creation

        Args:
            parent_data_folder (_type_, optional): _description_. Defaults to None.
            raw_data_folder (_type_, optional): _description_. Defaults to None.
            zipped_folder (_type_, optional): _description_. Defaults to None.
            excel_file (_type_, optional): _description_. Defaults to None.
            re_create_excel (bool, optional): _description_. Defaults to True.
        """
        super().__init__(parent_data_folder=parent_data_folder,
                raw_data_folder=raw_data_folder,
                zipped_folder=zipped_folder, )

        self.report_folder = PRPOJECT_PATH
        if not excel_file_prefix:
            excel_file_prefix="EDA_Reports"

        if not tmp_png_name_only:
            tmp_png_name_only = "tmp_png.png"
        self.tmp_png = self.report_folder + f"/{tmp_png_name_only}"

        if not excel_file:
            excel_file = self.report_folder + f"/reports/{excel_file_prefix}{datetime.strftime(datetime.now(), '%Y%m%d%H')}.xlsx"
        self.excel_file = excel_file

        #self.tmp_png = self.report_folder + f"/reports/tmp_{datetime.strftime(datetime.now(), '%Y%m%d%H')}.png"

        self.prepare_excel_report(re_create_excel=re_create_excel)

    # Excel
    def prepare_excel_report(self, re_create_excel=False):
        """
        Prepare Excel File for eda retports
        Args:
            re_create_excel (bool, optional): _description_. Defaults to False.
        """
        if re_create_excel or (not os.path.exists(self.excel_file)):
            if  os.path.exists(self.excel_file):
                os.remove(self.excel_file)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Notes"
            ws["A1"] = "About EDA"

            ws["A3"] = "File Locations:"
            n_row, n_col = 3, 1
            for k,v in self.file_folders.items():
                cell = ws.cell(n_row+1, n_col)
                cell.value = k
                cell = ws.cell(n_row+1, n_col+1)
                cell.value = v
                n_row += 1
            wb.save(self.excel_file)
            wb.close()

    # Excel
    def excel_add_data(self, data, sn, start_col, start_row, title):
        """
        Add data frame to excel sheet

        Args:
            data (_type_): can be data frame, np array or list
            sn (_type_): _description_
            start_col (_type_): _description_
            start_row (_type_): _description_
            title (_type_): _description_
        """
        wb = load_workbook(self.excel_file)
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            ws.title=sn
        else:
            ws = wb[sn]
        ws.cell(column=start_col,row=start_row,value=title)

        cols = list(data)
        start_row += 1

        if isinstance(data, pd.DataFrame):
            for i, col in enumerate(cols):
                ws.cell(column=start_col+i, row=start_row, value=col)

            start_row += 1
            vals = data.values
            nrow, ncol = vals.shape

            for r in range(nrow):
                for c in range(ncol):
                    ws.cell(column=start_col+c, row=start_row+r, value=vals[r,c])
        elif isinstance(data, list):
            for i in range(len(data)):
                ws.cell(column=start_col+i, row=start_row, value=data[i])
        elif isinstance(data, dict):
            ws.cell(column=start_col, row=start_row, value="Key")
            ws.cell(column=start_col+1, row=start_row, value="Value")

            start_row += 1
            for k,v in data.items():
                ws.cell(column=start_col, row=start_row, value=k)
                ws.cell(column=start_col+1, row=start_row, value=v)
                start_row += 1
        elif isinstance(data, np.ndarray):
            nd = len(data.shape)
            if nd == 1:
                data = np.expand_dims(data, axis=0)

            data = data.reshape(-1, 2)
            ii, jj = data.shape

            for i in range(ii):
                for j in range(jj):
                    ws.cell(column=start_col+jj, row=start_row+i, value=vals[i,j])
                    start_row += 1
        else:
            for i, v in enumerate(data):
                ws.cell(column=start_col, row=start_row+i, value=v)
                start_row += 1

        wb.save(self.excel_file)
        wb.close()

    # Excel
    def excel_chart(self, sn, title, start_row=2):
        """
        Load chart to Excel

        Args:
            sn (_type_): _description_
            title (_type_): _description_
            start_row (int, optional): _description_. Defaults to 2.
        """
        wb = load_workbook(self.excel_file)
        if sn not in wb.sheetnames:
            ws = wb.create_sheet(sn)
            ws.title=sn
        else:
            ws = wb[sn]
        ws[f"A{start_row}"] = title

        img = openpyxl.drawing.image.Image(self.tmp_png)
        img.anchor = f"A{start_row+2}"

        ws.add_image(img)

        wb.save(self.excel_file)
        wb.close()
